#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Genera ihpix_seed_data.json desde el archivo Excel de IHP-IX.

Uso:
    python generate_seed.py path/to/All_Priority_Areas_Reporting.xlsx [output.json]

Requiere: openpyxl
"""
import json
import sys
import os
import datetime


def _safe_int(val):
    if val is None:
        return 0
    try:
        return int(val)
    except (ValueError, TypeError):
        return 0


def _safe_str(val):
    if val is None:
        return u''
    if isinstance(val, datetime.datetime):
        return val.isoformat()
    return str(val).strip()


def _pa_key(full_name):
    """Convierte 'Priority Area 1: ...' → 'PA1'."""
    if not full_name:
        return u''
    s = full_name.strip()
    for i in range(1, 6):
        if s.startswith('Priority Area {}'.format(i)):
            return 'PA{}'.format(i)
    return s


def _output_key(output_text):
    """Extrae el código del output, e.g. '1.3. Research on...' → '1.3'."""
    if not output_text:
        return u''
    s = output_text.strip()
    # Tomar hasta el primer punto seguido de espacio después del número
    parts = s.split('.')
    if len(parts) >= 2:
        try:
            int(parts[0])
            int(parts[1].split(' ')[0].split('.')[0])
            return '{}.{}'.format(parts[0], parts[1].split(' ')[0].split('.')[0])
        except (ValueError, IndexError):
            pass
    return s


def generate_seed(excel_path, output_path=None):
    try:
        import openpyxl
    except ImportError:
        print('Error: openpyxl requerido. pip install openpyxl')
        sys.exit(1)

    if output_path is None:
        output_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            '..', 'data', 'ihpix_seed_data.json'
        )

    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # ── Actividades ──
    ws = wb['All Priority Areas']
    activities = []
    for r in range(2, ws.max_row + 1):
        title = _safe_str(ws.cell(r, 4).value)  # D: Name
        if not title:
            continue
        pa_full = _safe_str(ws.cell(r, 1).value)
        output_full = _safe_str(ws.cell(r, 2).value)

        activity = {
            'title': title,
            'priority_area': _pa_key(pa_full),
            'priority_area_full': pa_full,
            'output': _output_key(output_full),
            'output_full': output_full,
            'description': _safe_str(ws.cell(r, 5).value),
            'key_activity': _safe_str(ws.cell(r, 3).value),
            'outcomes': _safe_str(ws.cell(r, 6).value),
            'biennium': _safe_str(ws.cell(r, 7).value),
            'end_date': _safe_str(ws.cell(r, 8).value),
            'contact_name': _safe_str(ws.cell(r, 9).value),
            'contact_email': _safe_str(ws.cell(r, 10).value),
            'institution_type': _safe_str(ws.cell(r, 11).value),
            'institution': _safe_str(ws.cell(r, 12).value),
            'partners': _safe_str(ws.cell(r, 13).value),
            'unesco_participation': _safe_str(ws.cell(r, 14).value),
            'flagships': _safe_str(ws.cell(r, 15).value),
            'regions': _safe_str(ws.cell(r, 16).value),
            'member_states': _safe_str(ws.cell(r, 17).value),
            'knowledge_product_type': _safe_str(ws.cell(r, 18).value),
            'knowledge_product_type_other': _safe_str(ws.cell(r, 19).value),
            'num_knowledge_products': _safe_int(ws.cell(r, 20).value),
            'scientific_product_type': _safe_str(ws.cell(r, 21).value),
            'num_scientific_products': _safe_int(ws.cell(r, 22).value),
            'training_type': _safe_str(ws.cell(r, 23).value),
            'num_training_materials': _safe_int(ws.cell(r, 24).value),
            'num_curricula': _safe_int(ws.cell(r, 25).value),
            'num_transboundary_ms': _safe_int(ws.cell(r, 26).value),
            'knowledge_activity_type': _safe_str(ws.cell(r, 27).value),
            'knowledge_activity_type_other': _safe_str(ws.cell(r, 28).value),
            'stakeholders_knowledge': _safe_int(ws.cell(r, 29).value),
            'stakeholders_knowledge_female': _safe_int(ws.cell(r, 30).value),
            'stakeholders_knowledge_youth': _safe_int(ws.cell(r, 31).value),
            'stakeholders_awareness': _safe_int(ws.cell(r, 32).value),
            'stakeholders_awareness_female': _safe_int(ws.cell(r, 33).value),
            'stakeholders_awareness_youth': _safe_int(ws.cell(r, 34).value),
            'num_stakeholder_groups': _safe_int(ws.cell(r, 35).value),
            'stakeholder_group_type': _safe_str(ws.cell(r, 36).value),
            'notes': _safe_str(ws.cell(r, 37).value),
            'cross_cutting_wg': _safe_str(ws.cell(r, 38).value),
            'synergies': _safe_str(ws.cell(r, 39).value),
            'supporting_member_state': _safe_str(ws.cell(r, 40).value),
            'original_timestamp': _safe_str(ws.cell(r, 41).value),
            'original_id': _safe_str(ws.cell(r, 42).value),
            'status': 'published',
        }
        # Usar el primer member state como country si no hay supporting
        country = activity['supporting_member_state']
        if not country or country.lower() in ('no', 'yes', ''):
            ms = activity['member_states']
            if ms.startswith('["') and '"]' in ms:
                try:
                    ms_list = json.loads(ms)
                    country = ms_list[0] if ms_list else u''
                except (json.JSONDecodeError, ValueError):
                    country = u''
            else:
                country = ms
        activity['country'] = country
        activities.append(activity)

    print('Actividades procesadas: {}'.format(len(activities)))

    # ── Country Summaries (desde hoja Summary) ──
    ws_sum = wb['Summary']
    country_summaries = []
    for r in range(2, ws_sum.max_row + 1):
        country = _safe_str(ws_sum.cell(r, 1).value)
        if not country:
            continue
        cs = {
            'country': country,
            'latitude': float(ws_sum.cell(r, 2).value or 0),
            'longitude': float(ws_sum.cell(r, 3).value or 0),
            'region': _safe_str(ws_sum.cell(r, 4).value),
            'total_activities': _safe_int(ws_sum.cell(r, 5).value),
            'pa1_count': _safe_int(ws_sum.cell(r, 6).value),
            'pa2_count': _safe_int(ws_sum.cell(r, 7).value),
            'pa3_count': _safe_int(ws_sum.cell(r, 8).value),
            'pa4_count': _safe_int(ws_sum.cell(r, 9).value),
            'pa5_count': _safe_int(ws_sum.cell(r, 10).value),
            'transboundary_all': _safe_int(ws_sum.cell(r, 11).value),
            'transboundary_pa1': _safe_int(ws_sum.cell(r, 12).value),
            'transboundary_pa2': _safe_int(ws_sum.cell(r, 13).value),
            'transboundary_pa3': _safe_int(ws_sum.cell(r, 14).value),
            'transboundary_pa4': _safe_int(ws_sum.cell(r, 15).value),
            'transboundary_pa5': _safe_int(ws_sum.cell(r, 16).value),
            'supporting_all': _safe_int(ws_sum.cell(r, 17).value),
            'supporting_pa1': _safe_int(ws_sum.cell(r, 18).value),
            'supporting_pa2': _safe_int(ws_sum.cell(r, 19).value),
            'supporting_pa3': _safe_int(ws_sum.cell(r, 20).value),
            'supporting_pa4': _safe_int(ws_sum.cell(r, 21).value),
            'supporting_pa5': _safe_int(ws_sum.cell(r, 22).value),
        }
        country_summaries.append(cs)

    print('Países procesados: {}'.format(len(country_summaries)))

    # ── Enriquecer country summaries con PA output data y Flagships ──
    pa_sheets = {
        'Priority Area 1': ('pa1_outputs', ['1.1','1.2','1.3','1.4','1.5','1.6','1.7','1.8','1.9','1.10']),
        'Priority Area 2': ('pa2_outputs', ['2.1','2.2','2.3','2.4','2.5','2.6']),
        'Priority Area 3': ('pa3_outputs', ['3.1','3.2','3.3','3.4']),
        'Priority Area 4': ('pa4_outputs', ['4.1','4.2','4.3','4.4','4.5','4.6','4.7','4.8','4.9']),
        'Priority Area 5': ('pa5_outputs', ['5.1','5.2','5.3','5.4','5.5']),
    }

    # Crear diccionario por país para rápido acceso
    cs_by_country = {}
    for cs in country_summaries:
        cs_by_country[cs['country']] = cs
        cs['pa_output_data'] = {}
        cs['flagship_data'] = {}

    for sheet_name, (key, outputs) in pa_sheets.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws_pa = wb[sheet_name]
        for r in range(2, ws_pa.max_row + 1):
            country = _safe_str(ws_pa.cell(r, 1).value)
            if country not in cs_by_country:
                continue
            data = {}
            for i, out_code in enumerate(outputs):
                val = _safe_int(ws_pa.cell(r, 4 + i).value)
                if val:
                    data[out_code] = val
            if data:
                cs_by_country[country]['pa_output_data'][key] = data

    # Flagship data
    if 'Flagships' in wb.sheetnames:
        ws_fl = wb['Flagships']
        # Leer headers para nombres de flagship
        flagship_names = []
        for c in range(5, ws_fl.max_column + 1):
            h = _safe_str(ws_fl.cell(1, c).value)
            if h:
                flagship_names.append((c, h))
        for r in range(2, ws_fl.max_row + 1):
            country = _safe_str(ws_fl.cell(r, 1).value)
            if country not in cs_by_country:
                continue
            fl_data = {}
            for col, name in flagship_names:
                val = _safe_int(ws_fl.cell(r, col).value)
                if val:
                    fl_data[name] = val
            if fl_data:
                cs_by_country[country]['flagship_data'] = fl_data

    # Serializar JSON fields
    for cs in country_summaries:
        cs['pa_output_data'] = json.dumps(cs['pa_output_data'], ensure_ascii=False)
        cs['flagship_data'] = json.dumps(cs['flagship_data'], ensure_ascii=False)

    # ── Escribir JSON ──
    seed_data = {
        'version': '1.0',
        'generated_at': datetime.datetime.utcnow().isoformat(),
        'source': os.path.basename(excel_path),
        'activities': activities,
        'country_summaries': country_summaries,
    }

    output_path = os.path.abspath(output_path)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(seed_data, f, ensure_ascii=False, indent=2)

    print('Seed data escrito en: {}'.format(output_path))
    print('  Actividades: {}'.format(len(activities)))
    print('  Países: {}'.format(len(country_summaries)))
    return seed_data


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('Uso: python generate_seed.py <excel_path> [output.json]')
        sys.exit(1)
    excel_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else None
    generate_seed(excel_path, output_path)
